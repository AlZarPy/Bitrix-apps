from __future__ import annotations

import csv
import io
from typing import Any, Iterable

from openpyxl import load_workbook, Workbook


# --------- Утилиты нормализации ---------


def norm(s: str | None) -> str:
    return (s or "").strip().lower()


def norm_phone(raw: str | None) -> str:
    if not raw:
        return ""
    digits = "".join(ch for ch in raw if ch.isdigit())

    # минимальная нормализация, для поиска дублей
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    if len(digits) == 10:
        digits = "7" + digits
    return digits


def norm_email(raw: str | None) -> str:
    return (raw or "").strip().lower()


# --------- Парсинг файлов ---------


def _extract_row_common(row: dict[str, str]) -> dict[str, str]:
    """
    Приводим заголовки к единому виду.
    Базовый ожидаемый формат:
      имя, фамилия, номер телефона, почта, компания
    Дополнительно поддерживаем простые синонимы.
    row уже должен быть с ключами в lower-case.
    """
    mapping = {
        "first_name": ("имя", "first_name"),
        "last_name": ("фамилия", "last_name"),
        "phone": ("номер телефона", "телефон", "phone"),
        "email": ("почта", "email"),
        "company": ("компания", "company"),
    }

    def take(keys: Iterable[str]) -> str:
        for k in keys:
            v = row.get(k)
            if v:
                return str(v).strip()
        return ""

    return {
        "first_name": take(mapping["first_name"]),
        "last_name": take(mapping["last_name"]),
        "phone": take(mapping["phone"]),
        "email": take(mapping["email"]),
        "company": take(mapping["company"]),
    }


def parse_csv_file(file) -> list[dict[str, str]]:
    """
    Парсим CSV:
    - кодировка UTF-8-SIG,
    - разделитель ",",
    - нормализуем заголовки: strip + lower,
    - приводим к единому виду через _extract_row_common.
    """
    data = file.read()
    text = data.decode("utf-8-sig")

    if not text.strip():
        return []

    reader = csv.DictReader(io.StringIO(text), delimiter=",")

    rows: list[dict[str, str]] = []
    for raw_row in reader:
        if not raw_row:
            continue

        normalized_row: dict[str, str] = {}
        for k, v in raw_row.items():
            if not k:
                continue
            key = k.strip().lower()
            normalized_row[key] = (v or "").strip()

        rows.append(_extract_row_common(normalized_row))

    return rows


def parse_xlsx_file(file) -> list[dict[str, str]]:
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        headers = [str(h).strip().lower() for h in next(rows_iter)]
    except StopIteration:
        return []

    rows: list[dict[str, str]] = []
    for row_values in rows_iter:
        # пропускаем полностью пустые строки
        if not any(row_values):
            continue

        row: dict[str, str] = {}
        for h, v in zip(headers, row_values):
            if h:
                row[h] = "" if v is None else str(v)
        rows.append(_extract_row_common(row))
    return rows


def parse_uploaded_file(uploaded) -> list[dict[str, str]]:
    name = (uploaded.name or "").lower()
    if name.endswith(".csv"):
        return parse_csv_file(uploaded)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx_file(uploaded)
    raise ValueError("Поддерживаются только CSV и XLSX")


# --------- Работа с Bitrix24 ---------


def get_companies_map(but) -> dict[str, int]:
    """
    Словарь уже существующих компаний:
    normalized_title -> ID
    """
    items = but.call_list_method(
        "crm.company.list",
        fields={"select": ["ID", "TITLE"]},
    ) or []
    result: dict[str, int] = {}
    for c in items:
        title = norm(c.get("TITLE"))
        cid = c.get("ID")
        if title and cid:
            try:
                result[title] = int(cid)
            except (TypeError, ValueError):
                continue
    return result


def build_existing_contacts_index(but) -> set[tuple[str, str]]:
    """
    Индекс уже существующих контактов:
    ('phone', normalized) / ('email', normalized)
    Используем для отсечения дублей.
    """
    items = but.call_list_method(
        "crm.contact.list",
        fields={"select": ["ID", "PHONE", "EMAIL"]},
    ) or []

    index: set[tuple[str, str]] = set()

    for c in items:
        for p in c.get("PHONE", []) or []:
            np = norm_phone(p.get("VALUE"))
            if np:
                index.add(("phone", np))
        for e in c.get("EMAIL", []) or []:
            ne = norm_email(e.get("VALUE"))
            if ne:
                index.add(("email", ne))

    return index


def import_contacts(but, rows: list[dict[str, str]]) -> dict[str, int]:
    """
    Импорт контактов из уже распарсенных строк.
    - Матчим компанию по названию.
    - Не создаём дубли по телефону/почте.
    - Создание контактов отправляем в Bitrix батчами.
    """
    from urllib.parse import urlencode  # локальный import, чтобы не трогать заголовок файла

    companies = get_companies_map(but)
    existing_index = build_existing_contacts_index(but)

    created = 0
    skipped_duplicates = 0
    skipped_empty = 0

    seen_in_file: set[tuple[str, str]] = set()

    # команды для batch: key -> "crm.contact.add?fields[NAME]=..."
    batch_cmd: dict[str, str] = {}
    batch_size = 50  # лимит Bitrix для batch — до 50 команд за один вызов

    def flush_batch():
        nonlocal batch_cmd
        if not batch_cmd:
            return
        but.call_api_method("batch", params={"halt": 0, "cmd": batch_cmd})
        batch_cmd = {}

    for row in rows:
        fn = row.get("first_name", "").strip()
        ln = row.get("last_name", "").strip()
        phone_raw = row.get("phone", "").strip()
        email_raw = row.get("email", "").strip()
        company_raw = row.get("company", "").strip()

        # пустая строка (нет ни имени, ни фамилии)
        if not (fn or ln):
            skipped_empty += 1
            continue

        np = norm_phone(phone_raw)
        ne = norm_email(email_raw)

        # проверка дублей (в базе + внутри текущего файла)
        keys: list[tuple[str, str]] = []
        if np:
            keys.append(("phone", np))
        if ne:
            keys.append(("email", ne))

        if any(k in existing_index or k in seen_in_file for k in keys):
            skipped_duplicates += 1
            continue

        for k in keys:
            seen_in_file.add(k)

        payload: dict[str, Any] = {
            "NAME": fn,
            "LAST_NAME": ln,
        }

        if np:
            payload["PHONE"] = [{"VALUE": phone_raw, "VALUE_TYPE": "WORK"}]
        if ne:
            payload["EMAIL"] = [{"VALUE": email_raw, "VALUE_TYPE": "WORK"}]

        if company_raw:
            cid = companies.get(norm(company_raw))
            if cid:
                payload["COMPANY_ID"] = cid

        # добавляем команду в batch
        cmd_key = f"c{len(batch_cmd)}"
        batch_cmd[cmd_key] = "crm.contact.add?" + urlencode({"fields": payload})
        created += 1

        # если достигли размера батча — отправляем
        if len(batch_cmd) >= batch_size:
            flush_batch()

    # отправляем остаток, если есть
    flush_batch()

    return {
        "created": created,
        "skipped_duplicates": skipped_duplicates,
        "skipped_empty": skipped_empty,
    }



def _collect_contacts_for_export(
    but,
    date_from: str | None,
    date_to: str | None,
    company_filter: str | None = None,
) -> list[tuple[str, str, str, str, str]]:
    """
    Возвращает список кортежей:
    (имя, фамилия, телефон, почта, компания)
    — общая логика для CSV и XLSX экспорта.
    """
    filters: dict[str, Any] = {}
    if date_from:
        filters[">=DATE_CREATE"] = date_from + " 00:00:00"
    if date_to:
        filters["<=DATE_CREATE"] = date_to + " 23:59:59"

    contacts = but.call_list_method(
        "crm.contact.list",
        fields={
            "select": [
                "ID",
                "NAME",
                "LAST_NAME",
                "PHONE",
                "EMAIL",
                "COMPANY_ID",
                "COMPANY_TITLE",
            ],
            "filter": filters,
        },
    ) or []

    companies_map = get_companies_map(but)
    # id -> title (normalized)
    companies_by_id = {cid: title for title, cid in companies_map.items()}

    rows: list[tuple[str, str, str, str, str]] = []

    company_filter_norm = company_filter.lower() if company_filter else None

    for c in contacts:
        name = c.get("NAME") or ""
        last_name = c.get("LAST_NAME") or ""

        phone = ""
        if c.get("PHONE"):
            phone = c["PHONE"][0].get("VALUE", "") or ""

        email = ""
        if c.get("EMAIL"):
            email = c["EMAIL"][0].get("VALUE", "") or ""

        company_title = c.get("COMPANY_TITLE") or ""
        if not company_title:
            cid = c.get("COMPANY_ID")
            if cid:
                try:
                    company_title = companies_by_id.get(int(cid), "") or ""
                except (TypeError, ValueError):
                    pass

        # фильтр по компании (подстрока, регистр не важен)
        if company_filter_norm:
            if not company_title or company_filter_norm not in company_title.lower():
                continue

        rows.append((name, last_name, phone, email, company_title))

    return rows


def export_contacts_to_csv(
    but,
    date_from: str | None,
    date_to: str | None,
    company_filter: str | None = None,
) -> str:
    """
    Экспорт контактов в CSV:
    имя,фамилия,номер телефона,почта,компания
    """
    rows = _collect_contacts_for_export(but, date_from, date_to, company_filter)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["имя", "фамилия", "номер телефона", "почта", "компания"])

    for name, last_name, phone, email, company_title in rows:
        writer.writerow([name, last_name, phone, email, company_title])

    return output.getvalue()


def export_contacts_to_xlsx(
    but,
    date_from: str | None,
    date_to: str | None,
    company_filter: str | None = None,
) -> bytes:
    """
    Экспорт контактов в XLSX в том же формате:
    имя,фамилия,номер телефона,почта,компания
    """
    rows = _collect_contacts_for_export(but, date_from, date_to, company_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    ws.append(["имя", "фамилия", "номер телефона", "почта", "компания"])
    for row in rows:
        ws.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
